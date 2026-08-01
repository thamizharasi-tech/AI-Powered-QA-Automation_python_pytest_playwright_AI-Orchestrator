"""
Pipeline State Manager
======================
Persistent state store for the AI QA Orchestrator pipeline.

STORAGE
=======
All state is stored in two Excel workbooks (human-readable, auditable):

  pipeline_state.xlsx  — QA Pipeline State
  rtm_store.xlsx       — RTM Traceability

DEDUPLICATION (5 layers)
=========================
  1. Requirement  — SHA-256 hash → Cache sheet lookup → skip LLM if hit
  2. Test Cases   — TC-XXX-NNN regex → Test Cases sheet lookup
  3. Scripts      — file path → Scripts sheet lookup
  4. RTM rows     — REQ-ID::TC-ID composite key → RTM sheet lookup
  5. Excel data   — Column A check in DataAgent

FORCE RE-RUN
============
    state = PipelineStateManager(force=True)
    # or: FORCE_RERUN=1 python -m pytest ...
"""

import hashlib
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ─────────────────────────────────────────────────────────────────────────────
# Sheet column definitions
# ─────────────────────────────────────────────────────────────────────────────

_STATE_SHEETS: Dict[str, List[str]] = {
    "Requirements": [
        "Requirement Preview", "First Seen", "Last Run",
        "Run Count", "Agents Completed", "req_hash",
    ],
    "Agent Log": [
        "Requirement Preview", "Agent", "Status", "Saved At", "req_hash",
    ],
    "Test Cases": [
        "Test Case ID", "Test Scenario", "Test Type", "Pre-Conditions",
        "Test Steps", "Expected Result", "Actual Result", "Automation Status",
        "Priority", "RTM Link", "Defect ID", "Comments",
    ],
    "Scripts": [
        "Script Path", "Requirement Preview", "Created At",
    ],
    "Cache": [
        "req_hash", "agent_name", "output_text", "saved_at",
    ],
}

_RTM_SHEETS: Dict[str, List[str]] = {
    "RTM": [
        "Requirement ID", "User Story", "API / Screen Name", "Test Case ID",
        "Test Scenario", "Priority", "Test Type", "Automation Status",
        "Execution Status", "Defect ID", "Comments",
    ],
    "Key Scenarios": [
        "Key Scenario ID", "Linked Requirements", "req_hash", "Last Updated",
    ],
    "Risks": [
        "Risk ID", "Linked Requirements", "req_hash", "Last Updated",
    ],
    "Coverage": [
        "Requirement Preview", "Total Requirements", "Covered", "Not Covered",
        "Total Test Cases", "Automated", "Manual", "Total Risks",
        "Last Updated", "req_hash",
    ],
}

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") \
    if OPENPYXL_AVAILABLE else None
_HEADER_FONT = Font(color="FFFFFF", bold=True) if OPENPYXL_AVAILABLE else None

_COL_WIDTHS: Dict[str, int] = {
    "Requirement Preview": 55, "First Seen": 25, "Last Run": 25,
    "Run Count": 12, "Agents Completed": 50,
    "Agent": 25, "Status": 16, "Saved At": 25,
    "Test Case ID": 16, "Test Scenario": 50, "Test Type": 16,
    "Pre-Conditions": 40, "Test Steps": 60, "Expected Result": 45,
    "Actual Result": 35, "Automation Status": 18, "Priority": 12,
    "RTM Link": 20, "Defect ID": 14, "Comments": 30,
    "Script Path": 50, "Created At": 25,
    "Requirement ID": 16, "User Story": 28, "API / Screen Name": 35,
    "Execution Status": 18, "Key Scenario ID": 18,
    "Linked Requirements": 35, "Last Updated": 25, "Risk ID": 14,
    "Total Requirements": 20, "Covered": 12, "Not Covered": 14,
    "Total Test Cases": 18, "Automated": 14, "Manual": 12, "Total Risks": 14,
    "req_hash": 18, "agent_name": 25, "output_text": 80, "saved_at": 25,
}


# ─────────────────────────────────────────────────────────────────────────────
# Internal workbook helper
# ─────────────────────────────────────────────────────────────────────────────

class _Workbook:
    def __init__(self, path: Path, sheet_names: List[str]) -> None:
        self.path        = path
        self.sheet_names = sheet_names
        self._wb         = None

    def __enter__(self):
        self._wb = self._load_or_create()
        return self

    def __exit__(self, *args) -> None:
        if self._wb is not None:
            self._wb.save(self.path)
            self._wb = None

    def open(self):
        self._wb = self._load_or_create()
        return self

    def save(self) -> None:
        if self._wb is not None:
            self._wb.save(self.path)

    def close(self) -> None:
        self.save()
        self._wb = None

    def sheet(self, name: str):
        if self._wb is None:
            raise RuntimeError(f"Workbook '{self.path.name}' is not open.")
        return self._wb[name]

    def _load_or_create(self):
        all_headers = {**_STATE_SHEETS, **_RTM_SHEETS}
        if self.path.exists():
            wb = load_workbook(self.path)
            changed = False
            for name in self.sheet_names:
                if name not in wb.sheetnames:
                    ws = wb.create_sheet(name)
                    _write_header(ws, all_headers.get(name, []))
                    changed = True
            if changed:
                wb.save(self.path)
            return wb
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        for name in self.sheet_names:
            ws = wb.create_sheet(name)
            _write_header(ws, all_headers.get(name, []))
        wb.save(self.path)
        return wb


def _write_header(ws, headers: List[str]) -> None:
    if not headers:
        return
    ws.append(headers)
    if _HEADER_FILL and _HEADER_FONT:
        for cell in ws[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, header in enumerate(headers, start=1):
        width = _COL_WIDTHS.get(header, 20)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width


# ─────────────────────────────────────────────────────────────────────────────
# PipelineStateManager
# ─────────────────────────────────────────────────────────────────────────────

class PipelineStateManager:
    """
    Persistent state manager for the AI QA Orchestrator pipeline.
    Stores all state in two Excel workbooks.
    """

    STATE_DIR      = Path(__file__).parent / "pipeline_state"
    STATE_FILE     = STATE_DIR / "pipeline_state.xlsx"
    RTM_STORE_FILE = STATE_DIR / "rtm_store.xlsx"

    def __init__(self, force: bool = False) -> None:
        self.force = force
        self.STATE_DIR.mkdir(parents=True, exist_ok=True)

        if not OPENPYXL_AVAILABLE:
            print("  [StateManager] WARNING: openpyxl not installed.")
            return

        self._state_wb = _Workbook(self.STATE_FILE, list(_STATE_SHEETS.keys()))
        self._rtm_wb   = _Workbook(self.RTM_STORE_FILE, list(_RTM_SHEETS.keys()))

        self._state_wb.open()
        self._state_wb.close()
        self._rtm_wb.open()
        self._rtm_wb.close()

        print(f"  [StateManager] State : {self.STATE_FILE.name}")
        print(f"  [StateManager] RTM   : {self.RTM_STORE_FILE.name}")

    def __enter__(self):
        if OPENPYXL_AVAILABLE:
            self._state_wb.open()
            self._rtm_wb.open()
        return self

    def __exit__(self, *args) -> None:
        if OPENPYXL_AVAILABLE:
            self._state_wb.close()
            self._rtm_wb.close()

    # ── Requirement deduplication ─────────────────────────────────────────────

    def get_requirement_hash(self, requirement: str) -> str:
        normalised = " ".join(requirement.strip().split())
        return hashlib.sha256(normalised.encode("utf-8")).hexdigest()[:16]

    def is_requirement_processed(self, requirement: str) -> bool:
        if self.force or not OPENPYXL_AVAILABLE:
            return False
        req_hash = self.get_requirement_hash(requirement)
        return bool(self._get_cached(req_hash, "analysis"))

    def register_requirement(self, requirement: str) -> str:
        if not OPENPYXL_AVAILABLE:
            return self.get_requirement_hash(requirement)

        req_hash = self.get_requirement_hash(requirement)
        now      = datetime.now().isoformat()
        preview  = requirement.strip()[:120]

        in_context = self._state_wb._wb is not None
        if not in_context:
            self._state_wb.open()

        ws = self._state_wb.sheet("Requirements")
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[5].value == req_hash:
                rn        = row[0].row or 0
                _raw      = ws.cell(row=rn, column=4).value
                run_count = int(str(_raw or "0")) + 1
                ws.cell(row=rn, column=3).value = now
                ws.cell(row=rn, column=4).value = run_count
                print(f"  [StateManager] EXISTING requirement [{req_hash}] run #{run_count}")
                if not in_context:
                    self._state_wb.close()
                return req_hash

        ws.append([preview, now, now, 1, "", req_hash])
        print(f"  [StateManager] NEW requirement registered [{req_hash}]")
        if not in_context:
            self._state_wb.close()
        return req_hash

    # ── Agent output cache ────────────────────────────────────────────────────

    def get_cached_output(self, requirement: str, agent_name: str) -> Optional[str]:
        if self.force or not OPENPYXL_AVAILABLE:
            return None
        req_hash = self.get_requirement_hash(requirement)
        output   = self._get_cached(req_hash, agent_name)
        if output:
            print(f"  [StateManager] CACHE HIT  [{req_hash}] {agent_name}")
        return output

    def save_agent_output(self, requirement: str, agent_name: str, output: str) -> None:
        if not OPENPYXL_AVAILABLE:
            return

        req_hash = self.get_requirement_hash(requirement)
        preview  = requirement.strip()[:60]
        now      = datetime.now().isoformat()

        in_context = self._state_wb._wb is not None
        if not in_context:
            self._state_wb.open()

        ws_cache = self._state_wb.sheet("Cache")
        for row in ws_cache.iter_rows(min_row=2, values_only=False):
            if row[0].value == req_hash and row[1].value == agent_name:
                rn = int(row[0].row)
                ws_cache.cell(row=rn, column=3).value = output
                ws_cache.cell(row=rn, column=4).value = now
                break
        else:
            ws_cache.append([req_hash, agent_name, output, now])

        ws_log = self._state_wb.sheet("Agent Log")
        for row in ws_log.iter_rows(min_row=2, values_only=False):
            if row[4].value == req_hash and row[1].value == agent_name:
                rn = int(row[0].row)
                ws_log.cell(row=rn, column=3).value = "Completed"
                ws_log.cell(row=rn, column=4).value = now
                break
        else:
            ws_log.append([preview, agent_name, "Completed", now, req_hash])

        ws_req = self._state_wb.sheet("Requirements")
        for row in ws_req.iter_rows(min_row=2, values_only=False):
            if row[5].value == req_hash:
                rn       = int(row[0].row)
                existing = str(ws_req.cell(row=rn, column=5).value or "")
                agents   = [a.strip() for a in existing.split(",") if a.strip()]
                if agent_name not in agents:
                    agents.append(agent_name)
                ws_req.cell(row=rn, column=5).value = ", ".join(agents)
                break

        if not in_context:
            self._state_wb.close()
        print(f"  [StateManager] SAVED [{req_hash}] {agent_name}")

    # ── Test case registry ────────────────────────────────────────────────────

    def register_test_cases(self, requirement: str, test_cases_output: str) -> dict:
        if not OPENPYXL_AVAILABLE:
            return {}

        req_hash = self.get_requirement_hash(requirement)
        tc_ids   = list(set(re.findall(r'\bTC-[A-Z]+-\d{3}\b', test_cases_output)))
        results: Dict[str, str] = {}

        in_context = self._state_wb._wb is not None
        if not in_context:
            self._state_wb.open()

        ws       = self._state_wb.sheet("Test Cases")
        existing = {str(r[0]) for r in ws.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]}

        for tc_id in tc_ids:
            if tc_id in existing and not self.force:
                results[tc_id] = "DUPLICATE"
                print(f"  [StateManager] DUPLICATE TC: {tc_id}")
            else:
                details = self._extract_tc_details(test_cases_output, tc_id)
                ws.append([
                    tc_id,
                    details.get("scenario",       ""),
                    details.get("test_type",       "Positive"),
                    details.get("pre_conditions",  ""),
                    details.get("steps",           ""),
                    details.get("expected_result", ""),
                    "",
                    details.get("automation",      "Manual"),
                    details.get("priority",        "P3"),
                    details.get("rtm_link",        ""),
                    "NA",
                    "",
                ])
                existing.add(tc_id)
                results[tc_id] = "NEW"
                print(f"  [StateManager] NEW TC: {tc_id}")

        if not in_context:
            self._state_wb.close()
        return results

    def get_duplicate_tc_ids(self, test_cases_output: str) -> list:
        if not OPENPYXL_AVAILABLE:
            return []
        in_context = self._state_wb._wb is not None
        if not in_context:
            self._state_wb.open()
        ws       = self._state_wb.sheet("Test Cases")
        existing = {str(r[0]) for r in ws.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]}
        if not in_context:
            self._state_wb.close()
        tc_ids = re.findall(r'\bTC-[A-Z]+-\d{3}\b', test_cases_output)
        return [tc for tc in set(tc_ids) if tc in existing and not self.force]

    # ── Script registry ───────────────────────────────────────────────────────

    def register_script(self, requirement: str, script_path: str) -> str:
        if not OPENPYXL_AVAILABLE:
            return "NEW"

        preview = requirement.strip()[:60]
        now     = datetime.now().isoformat()

        in_context = self._state_wb._wb is not None
        if not in_context:
            self._state_wb.open()

        ws       = self._state_wb.sheet("Scripts")
        existing = {str(r[0]) for r in ws.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]}

        if script_path in existing and not self.force:
            print(f"  [StateManager] DUPLICATE SCRIPT: {script_path}")
            if not in_context:
                self._state_wb.close()
            return "DUPLICATE"

        ws.append([script_path, preview, now])
        if not in_context:
            self._state_wb.close()
        print(f"  [StateManager] NEW SCRIPT: {script_path}")
        return "NEW"

    def is_script_registered(self, script_path: str) -> bool:
        if self.force or not OPENPYXL_AVAILABLE:
            return False
        in_context = self._state_wb._wb is not None
        if not in_context:
            self._state_wb.open()
        ws       = self._state_wb.sheet("Scripts")
        existing = {str(r[0]) for r in ws.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]}
        if not in_context:
            self._state_wb.close()
        return script_path in existing

    # ── RTM store ─────────────────────────────────────────────────────────────

    def update_rtm_store(self, requirement: str, rtm_output: str) -> dict:
        if not OPENPYXL_AVAILABLE:
            return {"added": [], "skipped": []}

        req_hash = self.get_requirement_hash(requirement)
        now      = datetime.now().isoformat()
        summary: dict = {"added": [], "skipped": [], "req_hash": req_hash}

        req_ids  = list(set(re.findall(r'\bREQ-\d{3}\b',       rtm_output)))
        tc_ids   = list(set(re.findall(r'\bTC-[A-Z]+-\d{3}\b', rtm_output)))
        ks_ids   = list(set(re.findall(r'\bKS-\d{3}\b',        rtm_output)))
        risk_ids = list(set(re.findall(r'\bRISK-\d{3}\b',      rtm_output)))

        in_context = self._rtm_wb._wb is not None
        if not in_context:
            self._rtm_wb.open()

        ws_rtm = self._rtm_wb.sheet("RTM")
        existing_keys: set = {
            f"{r[0]}::{r[3]}"
            for r in ws_rtm.iter_rows(min_row=2, max_col=4, values_only=True)
            if r[0] and r[3]
        }

        for rtm_row in self._parse_rtm_rows(rtm_output, req_ids, tc_ids):
            key = f"{rtm_row['requirement_id']}::{rtm_row['test_case_id']}"
            if key in existing_keys and not self.force:
                summary["skipped"].append(key)
            else:
                ws_rtm.append([
                    rtm_row["requirement_id"],
                    rtm_row["user_story"],
                    rtm_row["api_name"],
                    rtm_row["test_case_id"],
                    rtm_row["test_scenario"],
                    rtm_row["priority"],
                    rtm_row["test_type"],
                    rtm_row["automation"],
                    rtm_row["status"],
                    rtm_row["defect_id"],
                    "",
                ])
                existing_keys.add(key)
                summary["added"].append(key)

        ws_ks = self._rtm_wb.sheet("Key Scenarios")
        existing_ks = {str(r[0]) for r in ws_ks.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]}
        for ks_id in ks_ids:
            linked = ",".join(self._linked_ids(rtm_output, ks_id, req_ids))
            if ks_id not in existing_ks or self.force:
                ws_ks.append([ks_id, linked, req_hash, now])
                existing_ks.add(ks_id)
                summary["added"].append(ks_id)
            else:
                summary["skipped"].append(ks_id)

        ws_risk = self._rtm_wb.sheet("Risks")
        existing_risks = {str(r[0]) for r in ws_risk.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]}
        for risk_id in risk_ids:
            linked = ",".join(self._linked_ids(rtm_output, risk_id, req_ids))
            if risk_id not in existing_risks or self.force:
                ws_risk.append([risk_id, linked, req_hash, now])
                existing_risks.add(risk_id)
                summary["added"].append(risk_id)
            else:
                summary["skipped"].append(risk_id)

        ws_cov       = self._rtm_wb.sheet("Coverage")
        total_reqs   = ws_rtm.max_row - 1
        total_tcs    = len(tc_ids)
        auto_count   = sum(
            1 for row in ws_rtm.iter_rows(min_row=2, values_only=True)
            if row[7] and str(row[7]).upper() in ("YES", "AUTOMATED", "PARTIAL")
        )
        manual_count = total_tcs - auto_count
        total_risks  = len(risk_ids)
        covered_reqs = len({
            str(row[0]) for row in ws_rtm.iter_rows(min_row=2, values_only=True)
            if row[0] and row[3] and str(row[3]) != "N/A"
        })
        not_covered  = max(0, total_reqs - covered_reqs)
        preview      = requirement.strip()[:60]

        for row in ws_cov.iter_rows(min_row=2, values_only=False):
            if row[9].value == req_hash:
                rn = int(row[0].row)
                for col, val in enumerate(
                    [None, total_reqs, covered_reqs, not_covered,
                     total_tcs, auto_count, manual_count, total_risks, now],
                    start=1
                ):
                    if val is not None:
                        ws_cov.cell(row=rn, column=col).value = val
                break
        else:
            ws_cov.append([
                preview, total_reqs, covered_reqs, not_covered,
                total_tcs, auto_count, manual_count, total_risks, now, req_hash,
            ])

        if not in_context:
            self._rtm_wb.close()
        return summary

    def get_rtm_coverage(self, requirement: str) -> dict:
        if not OPENPYXL_AVAILABLE:
            return {}
        req_hash   = self.get_requirement_hash(requirement)
        in_context = self._rtm_wb._wb is not None
        if not in_context:
            self._rtm_wb.open()
        ws = self._rtm_wb.sheet("Coverage")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[9] == req_hash:
                if not in_context:
                    self._rtm_wb.close()
                return {
                    "total_requirements": row[1],
                    "covered":            row[2],
                    "not_covered":        row[3],
                    "total_test_cases":   row[4],
                    "automated":          row[5],
                    "manual":             row[6],
                    "total_risks":        row[7],
                    "last_updated":       row[8],
                }
        if not in_context:
            self._rtm_wb.close()
        return {}

    def check_req_exists_in_rtm(self, req_id: str) -> bool:
        if not OPENPYXL_AVAILABLE:
            return False
        in_context = self._rtm_wb._wb is not None
        if not in_context:
            self._rtm_wb.open()
        ws       = self._rtm_wb.sheet("RTM")
        existing = {str(r[0]) for r in ws.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]}
        if not in_context:
            self._rtm_wb.close()
        return req_id in existing

    def check_tc_exists_in_rtm(self, tc_id: str) -> bool:
        if not OPENPYXL_AVAILABLE:
            return False
        in_context = self._rtm_wb._wb is not None
        if not in_context:
            self._rtm_wb.open()
        ws       = self._rtm_wb.sheet("RTM")
        existing = {str(r[3]) for r in ws.iter_rows(min_row=2, max_col=4, values_only=True) if r[3]}
        if not in_context:
            self._rtm_wb.close()
        return tc_id in existing

    # ── Summary & reporting ───────────────────────────────────────────────────

    def print_state_summary(self) -> None:
        if not OPENPYXL_AVAILABLE:
            print("  [StateManager] openpyxl not available.")
            return

        self._state_wb.open()
        self._rtm_wb.open()

        req_count = self._state_wb.sheet("Requirements").max_row - 1
        tc_count  = self._state_wb.sheet("Test Cases").max_row - 1
        scr_count = self._state_wb.sheet("Scripts").max_row - 1
        rtm_count = self._rtm_wb.sheet("RTM").max_row - 1

        print("\n" + "=" * 70)
        print("PIPELINE STATE SUMMARY")
        print("=" * 70)
        print(f"  State file       : {self.STATE_FILE}")
        print(f"  RTM file         : {self.RTM_STORE_FILE}")
        print(f"  Force re-run     : {self.force}")
        print(f"  Requirements     : {req_count} processed")
        print(f"  Test Cases       : {tc_count} registered")
        print(f"  Scripts          : {scr_count} registered")
        print(f"  RTM rows         : {rtm_count}")

        ws_req = self._state_wb.sheet("Requirements")
        if req_count > 0:
            print("\n  PROCESSED REQUIREMENTS:")
            for row in ws_req.iter_rows(min_row=2, values_only=True):
                if not row[5]:
                    continue
                print(f"    [{row[5]}] runs={row[3]} agents=[{row[4]}]")
                print(f"      Preview: {str(row[0] or '')[:60]}...")

        ws_tc = self._state_wb.sheet("Test Cases")
        if tc_count > 0:
            tc_ids = [str(r[0]) for r in ws_tc.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]]
            print(f"\n  REGISTERED TEST CASES ({len(tc_ids)}):")
            for tc_id in sorted(tc_ids):
                print(f"    {tc_id}")

        print("=" * 70 + "\n")
        self._state_wb.close()
        self._rtm_wb.close()

    def print_rtm_store_summary(self) -> None:
        if not OPENPYXL_AVAILABLE:
            return

        self._rtm_wb.open()
        ws_rtm  = self._rtm_wb.sheet("RTM")
        ws_ks   = self._rtm_wb.sheet("Key Scenarios")
        ws_risk = self._rtm_wb.sheet("Risks")
        ws_cov  = self._rtm_wb.sheet("Coverage")

        req_ids  = [str(r[0]) for r in ws_rtm.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]]
        ks_ids   = [str(r[0]) for r in ws_ks.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]]
        risk_ids = [str(r[0]) for r in ws_risk.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]]

        print("\n" + "=" * 70)
        print("RTM STORE SUMMARY")
        print("=" * 70)
        print(f"  RTM file         : {self.RTM_STORE_FILE}")
        print(f"  RTM rows         : {len(req_ids)}")
        print(f"  Key Scenarios    : {len(ks_ids)}")
        print(f"  Risks            : {len(risk_ids)}")

        if req_ids:
            print(f"\n  RTM ENTRIES ({len(req_ids)}):")
            for row in ws_rtm.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                print(f"    {row[0]} → TC: {row[3]} | Priority: {row[5]} | "
                      f"Auto: {row[7]} | Status: {row[8]}")

        for row in ws_cov.iter_rows(min_row=2, values_only=True):
            if not row[9]:
                continue
            print(f"\n  COVERAGE [{row[9]}]:")
            print(f"    Requirements : {row[2]}/{row[1]} covered")
            print(f"    Test Cases   : {row[5]}/{row[4]} automated")
            print(f"    Risks        : {row[7]} identified")
            print(f"    Last updated : {row[8]}")

        print("=" * 70 + "\n")
        self._rtm_wb.close()

    def list_processed_requirements(self) -> list:
        if not OPENPYXL_AVAILABLE:
            return []
        self._state_wb.open()
        ws_req   = self._state_wb.sheet("Requirements")
        ws_cache = self._state_wb.sheet("Cache")
        result   = []
        for row in ws_req.iter_rows(min_row=2, values_only=True):
            if not row[5]:
                continue
            req_hash = str(row[5])
            agents   = [str(r[1]) for r in ws_cache.iter_rows(min_row=2, values_only=True)
                        if r[0] == req_hash and r[1]]
            result.append({
                "hash":             req_hash,
                "text_preview":     str(row[0] or "")[:80] + "...",
                "first_seen":       row[1],
                "last_run":         row[2],
                "run_count":        row[3] or 0,
                "agents_completed": agents,
            })
        self._state_wb.close()
        return result

    def clear_state(self) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        for path in [self.STATE_FILE, self.RTM_STORE_FILE]:
            if path.exists():
                path.unlink()
        self._state_wb = _Workbook(self.STATE_FILE, list(_STATE_SHEETS.keys()))
        self._rtm_wb   = _Workbook(self.RTM_STORE_FILE, list(_RTM_SHEETS.keys()))
        self._state_wb.open()
        self._state_wb.close()
        self._rtm_wb.open()
        self._rtm_wb.close()
        print("  [StateManager] All state cleared.")

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _get_cached(self, req_hash: str, agent_name: str) -> Optional[str]:
        try:
            in_context = self._state_wb._wb is not None
            if not in_context:
                self._state_wb.open()
            ws = self._state_wb.sheet("Cache")
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == req_hash and row[1] == agent_name:
                    if not in_context:
                        self._state_wb.close()
                    return str(row[2]) if row[2] else None
            if not in_context:
                self._state_wb.close()
        except Exception:
            pass
        return None

    def _parse_rtm_rows(self, rtm_output: str, req_ids: list, tc_ids: list) -> List[dict]:
        rows: List[dict] = []
        table_lines = [
            line.strip() for line in rtm_output.split("\n")
            if line.strip().startswith("|") and "|" in line[1:]
        ]
        for line in table_lines:
            if "Requirement ID" in line or "---" in line or "===" in line:
                continue
            cells = [c.strip() for c in line.strip("|").split("|")]
            if len(cells) < 10:
                continue
            req_id_cell = cells[0].strip()
            tc_id_cell  = cells[3].strip()
            if not re.match(r'^(REQ|US|US0|REQ0)\w*', req_id_cell, re.IGNORECASE):
                continue
            rows.append({
                "requirement_id": req_id_cell,
                "user_story":     cells[1].strip() or "N/A",
                "api_name":       cells[2].strip() or "N/A",
                "test_case_id":   tc_id_cell or "N/A",
                "test_scenario":  cells[4].strip()[:80] or "N/A",
                "priority":       cells[5].strip() or "P3",
                "test_type":      cells[6].strip() or "Positive",
                "automation":     cells[7].strip() or "No",
                "status":         cells[8].strip() or "Not Executed",
                "defect_id":      cells[9].strip() or "NA",
            })
        if not rows:
            for req_id in req_ids:
                linked_tcs = self._linked_ids(rtm_output, req_id, tc_ids) or ["N/A"]
                for tc_id in linked_tcs:
                    rows.append({
                        "requirement_id": req_id,
                        "user_story":     "N/A",
                        "api_name":       "N/A",
                        "test_case_id":   tc_id,
                        "test_scenario":  f"Test case for {req_id}",
                        "priority":       "P2",
                        "test_type":      "Positive",
                        "automation":     "No",
                        "status":         "Not Executed",
                        "defect_id":      "NA",
                    })
        return rows

    def _extract_tc_details(self, test_cases_output: str, tc_id: str) -> dict:
        details: dict = {
            "scenario": "", "test_type": "", "pre_conditions": "",
            "steps": "", "expected_result": "", "automation": "Manual",
            "priority": "P3", "rtm_link": "",
        }
        lines    = test_cases_output.split("\n")
        tc_start = next((i for i, line in enumerate(lines) if tc_id in line), -1)
        if tc_start == -1:
            return details

        block = "\n".join(lines[tc_start: tc_start + 40])

        def _get(pattern: str, default: str = "") -> str:
            m = re.search(pattern, block, re.IGNORECASE | re.DOTALL)
            if m:
                val = m.group(1).strip().split("\n")[0].strip()
                return val[:200] if val else default
            return default

        title_m = re.search(r'(?:Title|Scenario)\s*[:\|]\s*(.+)', block, re.IGNORECASE)
        if title_m:
            details["scenario"] = title_m.group(1).strip()[:80]

        details["test_type"] = _get(
            r'Type\s*[:\|]\s*(FUNCTIONAL|NEGATIVE|BOUNDARY|SECURITY|INTEGRATION|'
            r'PERFORMANCE|REGRESSION|USABILITY)', "Positive"
        )

        prio_m = re.search(r'Priority\s*[:\|]\s*(P[1-4](?:-\w+)?)', block, re.IGNORECASE)
        if prio_m:
            p = prio_m.group(1).upper()
            details["priority"] = p[:2] if len(p) >= 2 else "P3"

        rtm_m = re.search(r'RTM\s*Link\s*[:\|]\s*([^\n]+)', block, re.IGNORECASE)
        if rtm_m:
            details["rtm_link"] = rtm_m.group(1).strip()[:50]

        auto_m = re.search(
            r'Automation\s*(?:Type)?\s*[:\|]\s*(YES|NO|PARTIAL|Automated|Manual|Partial)',
            block, re.IGNORECASE
        )
        if auto_m:
            val = auto_m.group(1).strip().upper()
            details["automation"] = (
                "Automated" if val in ("YES", "AUTOMATED")
                else "Partial" if val == "PARTIAL"
                else "Manual"
            )

        pre_m = re.search(
            r'Pre-?conditions?\s*[:\|]\s*([\s\S]+?)(?=Steps|Given|When|Expected|Post|─{5,}|$)',
            block, re.IGNORECASE
        )
        if pre_m:
            details["pre_conditions"] = " | ".join(
                line.strip().lstrip("-*").strip()
                for line in pre_m.group(1).strip().split("\n")
                if line.strip() and not line.strip().startswith("#")
            )[:300]

        steps_m = re.search(
            r'(?:Steps?\s*(?:\(Gherkin\))?\s*[:\|]?\s*)([\s\S]+?)(?=Expected|Post|─{5,}|$)',
            block, re.IGNORECASE
        )
        if steps_m:
            details["steps"] = " | ".join(
                line.strip()
                for line in steps_m.group(1).strip().split("\n")
                if line.strip() and not line.strip().startswith("#")
            )[:500]

        exp_m = re.search(r'Expected\s*Result\s*[:\|]\s*([^\n]+)', block, re.IGNORECASE)
        if exp_m:
            details["expected_result"] = exp_m.group(1).strip()[:200]

        return details

    def _linked_ids(self, text: str, anchor_id: str, candidates: list) -> list:
        lines       = text.split("\n")
        anchor_idxs = [i for i, line in enumerate(lines) if anchor_id in line]
        linked      = []
        for idx in anchor_idxs:
            context = " ".join(lines[max(0, idx - 3): idx + 4])
            for cid in candidates:
                if cid in context and cid not in linked:
                    linked.append(cid)
        return linked
