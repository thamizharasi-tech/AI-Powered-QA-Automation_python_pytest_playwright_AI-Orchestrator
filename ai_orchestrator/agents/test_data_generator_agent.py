"""
Test Data Agent
===============
Senior Test Data Engineer agent that generates comprehensive, structured test
data and writes it directly to testData/API_testData.xlsx via openpyxl.
"""

import json
import re
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class DataAgent:
    """Generate structured test data and write it to Excel for automation scripts."""

    _PROJECT_ROOT    = Path(__file__).resolve().parent.parent.parent
    DEFAULT_EXCEL_PATH = str(_PROJECT_ROOT / "testData" / "API_testData.xlsx")

    def __init__(self, llm, excel_path: str = "") -> None:
        self.llm        = llm
        self.excel_path = excel_path if excel_path else self.DEFAULT_EXCEL_PATH

    def generate_data(self, test_cases: str) -> str:
        prompt = self._build_prompt(test_cases)
        llm_output = self.llm.generate(prompt)
        self._try_write_to_excel(llm_output)
        return llm_output

    def write_excel_data(self, sheet_name: str, rows: list) -> dict:
        if not OPENPYXL_AVAILABLE:
            return {"error": "openpyxl not installed"}

        excel_path = Path(self.excel_path)
        excel_path.parent.mkdir(parents=True, exist_ok=True)

        wb = load_workbook(excel_path) if excel_path.exists() else Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            del wb["Sheet"]

        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

        existing_headers = [str(cell.value) for cell in ws[1] if cell.value] if ws.max_row >= 1 else []
        existing_row_names = {str(row[0]) for row in ws.iter_rows(min_row=2, max_col=1, values_only=True) if row[0]}

        written = skipped = 0
        for row_spec in rows:
            row_name = str(row_spec.get("row_name", ""))
            data     = row_spec.get("data", {})
            if not row_name:
                continue
            if row_name in existing_row_names:
                print(f"  [DataAgent] SKIP (already exists): {sheet_name}/{row_name}")
                skipped += 1
                continue
            new_headers = [h for h in data.keys() if h not in existing_headers]
            if new_headers:
                existing_headers.extend(new_headers)
                for col_idx, header in enumerate(existing_headers, start=1):
                    ws.cell(row=1, column=col_idx, value=header)
            next_row = ws.max_row + 1
            ws.cell(row=next_row, column=1, value=row_name)
            for col_idx, header in enumerate(existing_headers[1:], start=2):
                ws.cell(row=next_row, column=col_idx, value=data.get(header, ""))
            existing_row_names.add(row_name)
            written += 1
            print(f"  [DataAgent] WRITTEN: {sheet_name}/{row_name}")

        wb.save(excel_path)
        return {"written": written, "skipped": skipped, "sheet": sheet_name}

    def _try_write_to_excel(self, llm_output: str) -> None:
        """
        Parse LLM output and write test data to Excel.
        Supports two formats (both are tried):
          1. JSON blocks: ```json { "TC-XXX": { ... } } ```
          2. Markdown tables: ### SHEET N: `SheetName` | Col1 | Col2 | ...
        """
        if not OPENPYXL_AVAILABLE:
            return
        try:
            # ── Format 1: Fenced JSON blocks only (```json ... ```) ───────────
            # NOTE: We only try fenced blocks — NOT raw `{...}` patterns,
            # because those match too aggressively and set json_written=True
            # even when no real TC data was written.
            json_blocks = re.findall(r'```json\s*([\s\S]*?)```', llm_output)
            for block in json_blocks:
                try:
                    data = json.loads(block)
                    self._write_parsed_data(data)
                except json.JSONDecodeError:
                    continue

            # ── Format 2: Markdown tables (always run) ────────────────────────
            # Claude often returns Markdown tables instead of JSON.
            # We always run this parser regardless of whether JSON was found.
            self._parse_markdown_tables(llm_output)

        except Exception as e:
            print(f"  [DataAgent] Could not auto-write to Excel: {e}")

    def _parse_markdown_tables(self, llm_output: str) -> None:
        """
        Parse Markdown table sections from LLM output and write to Excel.

        Looks for patterns like:
          ### Sheet: SheetName
          | Row Identifier | col1 | col2 |
          |---|---|---|
          | TC_Feature_Valid | val1 | val2 |
        """
        if not OPENPYXL_AVAILABLE:
            return

        # Find sheet sections — matches patterns like:
        #   ### SHEET 1: `LeaveApplication` (KS-001 ...)
        #   ### Sheet: LeaveApplication
        #   ### SHEET: LeaveApplication
        sheet_pattern = re.compile(
            r'###\s+(?:SHEET|Sheet|sheet)\s*\d*\s*[:\-]\s*[`"\']?(\w+)[`"\']?',
            re.IGNORECASE
        )
        # Split output by sheet headings
        parts = sheet_pattern.split(llm_output)

        # parts = [text_before, sheet1_name, sheet1_content, sheet2_name, sheet2_content, ...]
        i = 1
        while i < len(parts) - 1:
            sheet_name = parts[i].strip()
            content    = parts[i + 1] if i + 1 < len(parts) else ""
            i += 2

            if not sheet_name or not content:
                continue

            # Extract the Markdown table from this section
            rows = self._extract_markdown_table_rows(content)
            if len(rows) < 2:
                continue  # Need at least header + 1 data row

            headers  = rows[0]
            data_rows = rows[1:]

            if not headers:
                continue

            # Build row specs for write_excel_data
            # Column A = row identifier (first column)
            rows_to_write = []
            for data_row in data_rows:
                if not data_row or not data_row[0]:
                    continue
                row_name = str(data_row[0]).strip()
                if not row_name or row_name.startswith('-'):
                    continue  # Skip separator rows
                row_data = {}
                for col_idx, header in enumerate(headers[1:], start=1):
                    if col_idx < len(data_row):
                        row_data[header] = data_row[col_idx].strip()
                rows_to_write.append({"row_name": row_name, "data": row_data})

            if rows_to_write:
                # Add Feature column to identify which sheet/feature each row belongs to
                for row_spec in rows_to_write:
                    row_spec["data"]["Feature"] = sheet_name

                # Write all data to Web_UI (single consolidated sheet)
                # This keeps all test data in one place for easy access via
                # read_api_data_from_excel("Web_UI", "TC_Feature_Variant")
                result = self.write_excel_data("Web_UI", rows_to_write)
                print(f"  [DataAgent] Markdown table → Web_UI (Feature={sheet_name}): "
                      f"{result.get('written', 0)} written, {result.get('skipped', 0)} skipped")

    def _extract_markdown_table_rows(self, content: str) -> list:
        """
        Extract rows from a Markdown table in the content.
        Returns list of lists (each inner list = one row's cells).
        """
        rows = []
        in_table = False
        for line in content.split('\n'):
            line = line.strip()
            if not line:
                if in_table:
                    break  # End of table
                continue
            if line.startswith('|'):
                in_table = True
                # Split by | and strip whitespace
                cells = [c.strip() for c in line.strip('|').split('|')]
                # Skip separator rows (---|---|---)
                if all(re.match(r'^[-:]+$', c.replace(' ', '')) for c in cells if c):
                    continue
                if cells:
                    rows.append(cells)
            elif in_table:
                break  # Non-table line after table started = end of table
        return rows

    def _write_parsed_data(self, data: dict) -> None:
        for key, value in data.items():
            if not isinstance(value, dict):
                continue
            if not (key.startswith("TC-") or key.startswith("KS-")):
                continue
            excel_mapping = value.get("excel_mapping", {})
            sheet_name    = excel_mapping.get("sheet_name", "TestData")
            rows_to_write = []
            for variant in value.get("positive_data", []):
                row_data = {"TestCaseID": key}
                row_data.update(variant.get("data", {}))
                row_data["ExpectedResult"] = variant.get("expected_result", "")
                row_data["DataType"] = "POSITIVE"
                rows_to_write.append({"row_name": f"{key}_{variant.get('variant','positive')}", "data": row_data})
            for variant in value.get("negative_data", []):
                row_data = {"TestCaseID": key}
                row_data.update(variant.get("data", {}))
                row_data["ExpectedResult"] = variant.get("expected_result", "")
                row_data["DataType"] = "NEGATIVE"
                rows_to_write.append({"row_name": f"{key}_{variant.get('variant','negative')}", "data": row_data})
            for variant in value.get("boundary_data", []):
                row_data = {"TestCaseID": key}
                row_data.update(variant.get("data", {}))
                row_data["ExpectedResult"] = variant.get("expected_result", "")
                row_data["DataType"] = "BOUNDARY"
                rows_to_write.append({"row_name": f"{key}_{variant.get('variant','boundary')}", "data": row_data})
            for variant in value.get("security_data", []):
                row_data = {"TestCaseID": key}
                row_data.update(variant.get("data", {}))
                row_data["ExpectedResult"] = variant.get("expected_result", "")
                row_data["DataType"] = "SECURITY"
                rows_to_write.append({"row_name": f"{key}_{variant.get('variant','security')}", "data": row_data})
            if rows_to_write:
                self.write_excel_data(sheet_name, rows_to_write)

    def _build_prompt(self, test_cases: str) -> str:
        return f"""
You are a Senior Test Data Engineer with 12+ years of experience in test data
management, data-driven testing, GDPR-compliant synthetic data generation, and
security testing.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
TEST CASES / KEY SCENARIOS:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{test_cases}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
EXCEL FRAMEWORK CONTRACT:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  read_api_data_from_excel(sheet_name, testcase_name) -> dict
  - Column A = row identifier (testcase_name)
  - Row 1 = column headers
  - Row identifier format: TC_[Feature]_[Variant]

REQUIRED OUTPUT SECTIONS:
## 1. EXCEL WORKBOOK STRUCTURE (table per sheet)
## 2. TEST DATA JSON (for programmatic Excel writing)

```json
{{
  "TC-[MODULE]-[NNN]": {{
    "tc_title": "[title]",
    "rtm_link": "[REQ-XXX]",
    "excel_mapping": {{
      "sheet_name": "[ActualFeatureName]",
      "row_name": "TC_[Feature]_[Variant]",
      "columns": ["[Field1]", "[Field2]", "DataType", "ExpectedResult"]
    }},
    "positive_data": [{{"variant": "valid", "data": {{"[Field1]": "[value]"}}, "expected_result": "[result]"}}],
    "negative_data": [{{"variant": "invalid", "data": {{"[Field1]": "[bad_value]"}}, "expected_result": "[error]"}}],
    "boundary_data": [{{"variant": "max", "data": {{"[Field1]": "[max_value]"}}, "expected_result": "[result]"}}],
    "security_data": [{{"variant": "sql_injection", "data": {{"[Field1]": "' OR '1'='1"}}, "expected_result": "Rejected"}}]
  }}
}}
```

## 3. AUTOMATION SCRIPT INTEGRATION GUIDE
## 4. DEDUPLICATION RULES
## 5. DATA DEPENDENCIES

OUTPUT RULES:
✅ Generate data for EVERY test case provided
✅ Use realistic synthetic data (not "test1", "foo", "bar")
✅ Include at least 1 positive, 2 negative, 2 boundary, 2 security variants
✅ Ensure all JSON is valid and properly escaped
❌ Do NOT use real personal information
❌ Do NOT hardcode data in automation scripts
"""
