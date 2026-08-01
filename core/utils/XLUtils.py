"""
XLUtils — Excel Test Data Utilities
=====================================
Provides helpers to read and write test data from/to the Excel workbook
(testData/API_testData.xlsx).

All public functions are pure functions (no `self` parameter) so they can
be called directly from test classes and module-level code without passing
a test-instance reference.

Usage:
    from core.utils.XLUtils import read_api_data_from_excel, read_cloud_env

    data = read_api_data_from_excel("Web_UI", "TC_Login_Valid")
    username = str(data.get("Username", ""))
"""

import glob
import os

import openpyxl

from core.e2e_testData import xlsx_apiTestData_file


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _get_latest_file(pattern: str) -> str:
    """
    Return the most recently created file matching the glob pattern.

    Raises
    ------
    FileNotFoundError — if no file matches the pattern
    """
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(
            f"Excel file not found: '{pattern}'. "
            "Run the DataAgent pipeline first to populate testData/API_testData.xlsx."
        )
    return max(files, key=os.path.getctime)


def _find_row_number(worksheet, name: str) -> int:
    """
    Search column A of the worksheet for a cell whose value equals `name`.

    Returns the 1-based row number of the first match.

    Raises
    ------
    ValueError — if `name` is not found in column A
    """
    for row in worksheet.iter_rows(min_col=1, max_col=1, values_only=False):
        cell = row[0]
        if cell.value == name:
            return int(cell.row)
    raise ValueError(
        f"Row '{name}' not found in sheet '{worksheet.title}'. "
        "Check the row identifier in Column A of the Excel file."
    )


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def read_api_data_from_excel(sheet_name: str, testcase_name: str) -> dict:
    """
    Read a named row from the Excel workbook as a key-value dictionary.

    Loads the workbook once, finds the row by name in Column A, then
    maps every column header (row 1) to the corresponding cell value.

    Parameters
    ----------
    sheet_name    : str — Excel sheet name (e.g. "Web_UI", "UserManagement")
    testcase_name : str — Row identifier in Column A (e.g. "TC_Login_Valid")

    Returns
    -------
    dict — {column_header: cell_value, ...}

    Raises
    ------
    FileNotFoundError — if the Excel file does not exist
    ValueError        — if the sheet or row is not found
    """
    excel_path = _get_latest_file(xlsx_apiTestData_file)
    workbook   = openpyxl.load_workbook(excel_path)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in '{excel_path}'. "
            f"Available sheets: {workbook.sheetnames}"
        )

    worksheet = workbook[sheet_name]
    row_number = _find_row_number(worksheet, testcase_name)
    col_count  = worksheet.max_column

    result: dict = {}
    for col_idx in range(1, col_count + 1):
        header = worksheet.cell(row=1,          column=col_idx).value
        value  = worksheet.cell(row=row_number, column=col_idx).value
        if header is not None:
            result[header] = value

    return result


def save_api_resp_in_excel(sheet_name: str, testcase_name: str, data: str) -> None:
    """
    Write an API response value into column 8 of the matching row.

    Parameters
    ----------
    sheet_name    : str — Excel sheet name
    testcase_name : str — Row identifier in Column A
    data          : str — Value to write into column 8 of that row
    """
    excel_path = _get_latest_file(xlsx_apiTestData_file)
    workbook   = openpyxl.load_workbook(excel_path)

    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in '{excel_path}'.")

    worksheet  = workbook[sheet_name]
    row_number = _find_row_number(worksheet, testcase_name)
    worksheet.cell(row=row_number, column=8).value = data
    workbook.save(excel_path)


def read_cloud_env() -> str:
    """
    Return the current test environment name (e.g. "QA", "UAT", "PROD").

    Reads from sheet "API_Common_Data", row "Test Env Name", column "Value".

    Returns
    -------
    str — environment name, or "QA" if not found
    """
    try:
        data = read_api_data_from_excel("API_Common_Data", "Test Env Name")
        return str(data.get("Value", "QA"))
    except Exception:
        return "QA"


# ─────────────────────────────────────────────────────────────────────────────
# Parametrize helper — read multiple Excel rows for @pytest.mark.parametrize
# ─────────────────────────────────────────────────────────────────────────────

def get_excel_rows(sheet_name: str, row_names: list) -> list:
    """
    Read multiple named rows from Excel for use with @pytest.mark.parametrize.

    Calls read_api_data_from_excel() for each row name and returns a list
    of dicts. Missing rows fall back to {"_row_name": row_name} so tests
    still run with defaults.

    Parameters
    ----------
    sheet_name : str  — Excel sheet name (e.g. "UserManagement")
    row_names  : list — Column A row identifiers to read

    Returns
    -------
    list[dict] — one dict per row; each dict has "_row_name" injected for
                 use as the parametrize test ID

    Example:
        from core.utils.XLUtils import get_excel_rows

        @pytest.mark.parametrize("data", get_excel_rows("UserManagement", [
            "UserManagement_Create_User",
            "UserManagement_Create_User_Admin",
        ]), ids=lambda d: d.get("_row_name", "unknown"))
        def test_create_user(self, data):
            username = data.get("username", "fallback")
    """
    rows = []
    for row_name in row_names:
        try:
            row_data = read_api_data_from_excel(sheet_name, row_name)
            row_data["_row_name"] = row_name
            rows.append(row_data)
        except Exception:
            rows.append({"_row_name": row_name})
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Legacy shim — keeps backward compatibility with old call sites that pass
# `self` as the first argument (e.g. read_api_data_from_excel(self, sheet, row))
# These wrappers accept and silently discard the `self` argument.
# New code should call the pure functions above directly.
# ─────────────────────────────────────────────────────────────────────────────

def get_rowNo_from_excel(workbook_path: str, sheetname: str, name: str):
    """
    Return (workbook, worksheet, row_number) for the given row name.

    Raises ValueError if the row is not found (no silent UnboundLocalError).
    """
    excel_path = _get_latest_file(workbook_path)
    wb         = openpyxl.load_workbook(excel_path)
    ws         = wb[sheetname]
    row_no     = _find_row_number(ws, name)
    return wb, ws, row_no
